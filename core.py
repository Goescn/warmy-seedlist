"""
Warmy Seed List — 核心业务逻辑
可被 server.py (API) 和 send_warmy_seedlist.py (CLI) 调用

所有敏感凭据优先读取环境变量，其次使用默认值（本地开发用）
"""
import json, os, re, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formataddr
from email import encoders
from openpyxl import Workbook
from datetime import datetime

import requests

# ─── 环境变量或默认值 ──────────────────────────────────

def env_or(key, default=""):
    return os.environ.get(key, default)

# Warmy
WARMY_API_KEY    = env_or("WARMY_API_KEY",    "Y5VMMbYPUgqcv3FLa7xvqZJtEPuVu4idVZyme1y7hYKRgE2P2WBqQ549Vsjj8QcWsMHs2U4j8knghaJg4eGWUXuxpfwbX3gT9NoH")
WARMY_HOLDER_UID = env_or("WARMY_HOLDER_UID", "a66a9a755fe16f24fcb99dc8b5f25a50")

# 飞书
FEISHU_APP_ID     = env_or("FEISHU_APP_ID",     "cli_a97a899f97785cc0")
FEISHU_APP_SECRET = env_or("FEISHU_APP_SECRET", "lPu3zwKOe4AUFBzrHQ1OmdiMpWTXPjnr")
FEISHU_WEBHOOK_URL= env_or("FEISHU_WEBHOOK_URL","https://open.feishu.cn/open-apis/bot/v2/hook/80671778-a115-42be-8677-be5b895c81ab")

# SMTP (exmail)
SMTP_HOST = env_or("SMTP_HOST", "smtp.exmail.qq.com")
SMTP_PORT = int(env_or("SMTP_PORT", "465"))
SMTP_USER = env_or("SMTP_USER", "kai.liu@expertsender.cn")
SMTP_PASS = env_or("SMTP_PASS", "CvyGqw4NYr6R8fd2")

# 默认目标
DEFAULT_TARGET_OPENID = env_or("TARGET_OPENID", "ou_1a04df09dbec2e78b7d2c9fd82bab012")
DEFAULT_TARGET_EMAIL  = env_or("TARGET_EMAIL",  "bing.liu@expertsender.cn")

# ─── 工具函数 ──────────────────────────────────────────

def sanitize_filename(name):
    safe = re.sub(r'[^a-zA-Z0-9._-]', '_', name)
    return safe.strip('_')


# ─── Warmy API ─────────────────────────────────────────

def fetch_split_info(split_id):
    resp = requests.get(
        f"https://api.warmy.io/api/v2/users_splits/{split_id}",
        headers={
            "Authorization": f"Bearer {WARMY_API_KEY}",
            "Holder-Uid": WARMY_HOLDER_UID,
            "Accept": "application/json"
        },
        timeout=15
    )
    resp.raise_for_status()
    return resp.json()


def fetch_seedlist(split_id):
    """分页拉取种子列表，返回 (email_list, pagination_dict)"""
    emails = []
    page = 1
    total_pages = 1
    while page <= total_pages:
        resp = requests.get(
            f"https://api.warmy.io/api/v2/users_splits/{split_id}/seedlist_emails",
            params={"page": page},
            headers={
                "Authorization": f"Bearer {WARMY_API_KEY}",
                "Holder-Uid": WARMY_HOLDER_UID,
                "Accept": "application/json"
            },
            timeout=30
        )
        resp.raise_for_status()
        data = resp.json()
        page_emails = data.get("items", data.get("data", []))
        if not page_emails:
            break
        emails.extend(page_emails)
        pag = data.get("pagination", data.get("meta", {}))
        total_pages = pag.get("total_pages", 1)
        page += 1
    return emails, {"total_count": len(emails), "total_pages": total_pages}


# ─── Excel ─────────────────────────────────────────────

def create_excel(emails, sender_name, output_dir=None):
    """生成 Excel，返回 (count, filepath, filename)"""
    today = datetime.now().strftime('%Y%m%d')
    label = sanitize_filename(sender_name)
    filename = f"{label}_seedlist_{today}.xlsx"
    out = output_dir or os.path.expanduser("~/Desktop")
    filepath = os.path.join(out, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Seed List Emails"
    ws["A1"] = "No."
    ws["B1"] = "Email Address"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    for i, email in enumerate(emails, 1):
        ws[f"A{i+1}"] = i
        ws[f"B{i+1}"] = email
    wb.save(filepath)

    return len(emails), filepath, filename


# ─── Feishu Drive ──────────────────────────────────────

def get_tenant_token():
    resp = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET},
        timeout=10
    )
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"获取 tenant token 失败: {data.get('msg')}")
    return data["tenant_access_token"]


def upload_file_to_drive(token, filepath, filename):
    with open(filepath, "rb") as f:
        resp = requests.post(
            "https://open.feishu.cn/open-apis/drive/v1/files/upload_all",
            headers={"Authorization": f"Bearer {token}"},
            data={
                "file_name": filename,
                "parent_type": "explorer",
                "size": str(os.path.getsize(filepath))
            },
            files={"file": (filename, f,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30
        )
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"云盘上传失败: {data.get('msg')}")
    ft = data["data"]["file_token"]
    return ft, f"https://icnxvmdqjlv8.feishu.cn/file/{ft}"


def share_file_with_user(token, file_token, openid):
    resp = requests.post(
        f"https://open.feishu.cn/open-apis/drive/v1/permissions/{file_token}/members",
        params={"type": "file", "need_notification": "false"},
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"member_type": "openid", "member_id": openid, "perm": "full_access"},
        timeout=10
    )
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"共享文件失败: {data.get('msg')}")


# ─── SMTP ──────────────────────────────────────────────

def send_email_smtp(filepath, filename, to_email):
    if not to_email:
        return None
    msg = MIMEMultipart()
    msg["From"] = formataddr(("Warmy Seed List Bot", SMTP_USER))
    msg["To"] = to_email
    msg["Subject"] = f"Warmy Seed List - {filename}"
    msg.attach(MIMEText(
        f"📊 Warmy Seed List 自动生成\n文件: {filename}\n"
        f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n请查收附件。",
        "plain", "utf-8"
    ))
    with open(filepath, "rb") as f:
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)
    return to_email


# ─── Feishu Webhook Card ───────────────────────────────

def send_webhook_card(results):
    today = datetime.now().strftime('%Y-%m-%d')
    fields = []
    for r in results:
        fields.append({
            "is_short": True,
            "text": {
                "tag": "lark_md",
                "content": f"**{r['sender']}**\n📧 {r['count']} 个邮箱"
            }
        })
    elements = [{"tag": "div", "fields": fields}, {"tag": "hr"}]
    for i, r in enumerate(results, 1):
        line = f"**{i}. {r['sender']}**\n📧 {r['count']} 个邮箱\n"
        if r.get('url'):
            line += f"🔗 [下载 {r['filename']}]({r['url']})"
        if r.get('email_sent'):
            line += f"\n📬 已发送至 {r['email_sent']}"
        elements.append({"tag": "div", "text": {"tag": "lark_md", "content": line}})
    elements.append({
        "tag": "note",
        "elements": [{"tag": "plain_text",
                      "content": f"⏰ {today} · Warmy Seed List 自动同步"}]
    })
    card = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": "📊 Warmy Seed List 周报"},
            "template": "blue"
        },
        "elements": elements
    }
    resp = requests.post(FEISHU_WEBHOOK_URL,
                         json={"msg_type": "interactive", "card": card},
                         timeout=15)
    return resp


# ─── 主执行函数 ────────────────────────────────────────

def run(senders, target_email=None, target_openid=None, output_dir=None):
    """
    完整执行流程: fetch → Excel → 上传飞书 → 发送邮件 → webhook 卡片
    senders: [{"split_id": str, "name": str}, ...]
    返回: {"success": bool, "results": [...], "email_sent_to": str|None, "error": str|None}
    """
    if target_openid is None:
        target_openid = DEFAULT_TARGET_OPENID
    if target_email is None:
        target_email = DEFAULT_TARGET_EMAIL
    output_dir = output_dir or os.path.expanduser("~/Desktop")

    results = []
    for s in senders:
        sid, sname = s["split_id"], s["name"]
        print(f"── {sname} (Split {sid}) ──")
        try:
            info = fetch_split_info(sid)
            print(f"  Info: {info.get('provider')}, {info.get('group_name')}")
        except Exception as e:
            print(f"  ❌ {e}")
            results.append({"sender": sname, "error": str(e), "count": 0})
            continue

        try:
            emails, pag = fetch_seedlist(sid)
            print(f"  ✅ {pag['total_count']} emails")
        except Exception as e:
            print(f"  ❌ {e}")
            results.append({"sender": sname, "error": str(e), "count": 0})
            continue

        try:
            cnt, fp, fn = create_excel(emails, sname, output_dir)
            print(f"  📝 {fn} — {cnt} emails")
        except Exception as e:
            print(f"  ❌ {e}")
            results.append({"sender": sname, "error": str(e), "count": 0})
            continue

        results.append({"sender": sname, "split_id": sid, "count": cnt,
                        "filepath": fp, "filename": fn, "emails": emails})

    if not any(r.get("count", 0) for r in results):
        return {"success": False, "results": results, "error": "没有成功拉取的数据"}

    # 上传飞书
    try:
        token = get_tenant_token()
        for r in results:
            if r.get("count", 0) == 0:
                continue
            ft, url = upload_file_to_drive(token, r["filepath"], r["filename"])
            share_file_with_user(token, ft, target_openid)
            r["url"] = url
    except Exception as e:
        print(f"  ⚠️ 飞书上传跳过: {e}")
        for r in results:
            r["url"] = None

    # 发送邮件
    sent_to = None
    if target_email:
        for r in results:
            if r.get("count", 0) == 0:
                continue
            try:
                sent = send_email_smtp(r["filepath"], r["filename"], target_email)
                r["email_sent"] = sent
                if sent:
                    sent_to = sent
            except Exception as e:
                print(f"  ❌ 邮件失败: {e}")

    # Webhook 卡片
    try:
        send_webhook_card(results)
    except Exception as e:
        print(f"  ⚠️ Webhook 跳过: {e}")

    return {"success": True, "results": results, "email_sent_to": sent_to}
